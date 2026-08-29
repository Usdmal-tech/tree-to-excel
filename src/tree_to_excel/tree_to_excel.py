#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Console tool to convert tree command output to Excel.
Supports flat, merged, and merged_full modes.
For full documentation, see README.md.
For use: python tree_to_excel.py -i <tree_file> [-m <mode>] [-o <output_file>] [-e <encoding>]
"""

__version__ = "1.0.0-cli"

import os
import sys
import argparse
import chardet

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------- Constants -----------------------------
ENCODING_CONFIDENCE_THRESHOLD = 0.85
INDENT_STEP = 4
LEVEL_COLORS = [
    "D9E1F2",  # L1
    "E2EFDA",  # L2
    "EFE9D8",  # L3
    "FCE4D6",  # L4
    "E4DFEC",  # L5
    "DAEEF3",  # L6
]

# ------------------ File reading with encoding -------------------
def read_tree_lines(filepath: str, forced_encoding: str = None) -> list[str]:
    """
    Reads the tree file with encoding detection or forced encoding.
    
    Args:
        filepath: Path to the tree output file.
        forced_encoding: If provided, uses this encoding unconditionally.
    
    Returns:
        List of lines (without trailing newlines).
    
    Raises:
        ValueError: If decoding fails even with forced encoding.
    
    Complexity: O(N) – reads entire file once.
    Safety: Handles UnicodeDecodeError gracefully and suggests manual encoding.
    """
    with open(filepath, "rb") as f:
        raw = f.read()

    if forced_encoding:
        encoding = forced_encoding
    else:
        detected = chardet.detect(raw)
        encoding = detected.get("encoding")
        confidence = detected.get("confidence", 0.0)

        if encoding is None or confidence < ENCODING_CONFIDENCE_THRESHOLD:
            encoding = "utf-8"
            print(
                f"Warning: Could not reliably detect encoding (confidence={confidence:.2f}). "
                f"Using 'utf-8' as fallback. If characters are garbled, rerun with --encoding.",
                file=sys.stderr
            )

    try:
        text = raw.decode(encoding)
    except UnicodeDecodeError as e:
        raise ValueError(
            f"Failed to decode file with {encoding}: {e}. "
            "Please specify correct encoding with --encoding."
        )

    return text.splitlines()

# -------------------------- Tree parser---------------------------
def parse_tree_file(filepath: str, encoding: str = None):
    """
    Parses tree output file (supports formats with and without the /A parameter).
    
    Args:
        filepath: Path to the tree file.
        encoding: Optional forced encoding (overrides auto-detection).
    
    Returns:
        (items, folder_count, file_count)
    """
    lines = read_tree_lines(filepath, encoding) # lines are now Unicode strings

    # --------------------- Find root path ---------------------
    root_line = None
    start_idx = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if len(stripped) > 2 and stripped[1] == ":" and "\\" in stripped:
            root_line = stripped
            start_idx = i + 1
            break
    if root_line is None:
        raise ValueError(r"Root path not found in format 'X:\...'")

    is_standard = False
    for line in lines[start_idx:start_idx+10]:
        if '├' in line or '└' in line:
            is_standard = True
            break

    # -------------------- Collect all items ---------------------
    raw_items = []  # (level, name, has_marker)
    markers = ["+---", "\\---", "├───", "└───"]

    for line in lines[start_idx:]:
        line = line.rstrip("\n\r")
        if not line.strip():
            continue

        # Find marker
        marker_pos = -1
        for marker in markers:
            p = line.find(marker)
            if p != -1:
                marker_pos = p
                break

        if marker_pos != -1:
            # Line with marker - folder (or empty folder)
            level = marker_pos // INDENT_STEP + 1
            name = line[marker_pos + INDENT_STEP:].strip()
            has_marker = True
        else:
            # Line without marker is file (or indented item without marker)
            # Remove all leading special characters (spaces, |, │, \, +, -)
            stripped = line.lstrip("│| \\+-")
            if not stripped:
                continue
            indent_len = len(line) - len(stripped)
            level = indent_len // INDENT_STEP
            if level == 0:
                level = 1
            name = stripped.strip()
            has_marker = False

        if not name:
            continue
        raw_items.append((level, name, has_marker))

    if not raw_items:
        return [], 0, 0

    # ---------- Determine is_dir (children + marker) ----------
    total_items = len(raw_items)
    items = []
    for idx, (level, name, has_marker) in enumerate(raw_items):
        # Check for children
        has_child = False
        for j in range(idx + 1, total_items):
            if raw_items[j][0] <= level:
                break
            if raw_items[j][0] > level:
                has_child = True
                break

        # Folder if has children OR (no children but had marker - empty folder)
        if is_standard:
            is_dir = has_child or (has_marker and not has_child and '.' not in name)
        else:
            # /A format: folder only by presence of children (marker not considered)
            is_dir = has_child

        items.append({
            "level": level,
            "name": name,
            "is_dir": is_dir,
            "ancestors": [],
            "path": None,
        })

    # ---------------- Build ancestors and path ----------------
    stack = []
    for item in items:
        # Remove from stack all items with level >= current
        while stack and stack[-1][0] >= item["level"]:
            stack.pop()

        # ancestors - folder names in stack
        item["ancestors"] = [s[1] for s in stack]

        # Build full path
        if item["is_dir"]:
            path_parts = [root_line] + [s[1] for s in stack] + [item["name"]]
        else:
            path_parts = [root_line] + [s[1] for s in stack] + [item["name"]]
        item["path"] = os.path.join(*path_parts)

        # If item is folder, add to stack
        if item["is_dir"]:
            stack.append((item["level"], item["name"]))

    # ---------------- Count folders and files -----------------
    folder_count = sum(1 for item in items if item["is_dir"])
    file_count = len(items) - folder_count

    return items, folder_count, file_count

# -------------------------- Save in Excel --------------------------
def save_to_excel(items: list[dict], output_file: str, mode: str, folder_count: int, file_count: int) -> None:
    """Saves the list of items to Excel with the selected mode."""
    if not items:
        return

    # --------------- Filter for flat and merged ---------------
    if mode in ("flat", "merged"):
        items = [item for item in items if item["is_dir"]]

    if not items:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Structure"

    max_level = max(item["level"] for item in items)

    headers = [f"L{i+1}" for i in range(max_level)] + ["Full path"]
    ws.append(headers)

    if mode == "flat":
        for item in items:
            row = [""] * (max_level + 1)
            if 1 <= item["level"] <= max_level:
                row[item["level"] - 1] = item["name"]
            row[-1] = item["path"]
            ws.append(row)
    else:
        # merged / merged_full
        for item in items:
            row = [""] * (max_level + 1)
            for level in range(1, item["level"] + 1):
                if level <= len(item["ancestors"]):
                    row[level - 1] = item["ancestors"][level - 1]
                elif level == item["level"]:
                    row[level - 1] = item["name"]
            row[-1] = item["path"]
            ws.append(row)

    # ------------------------ Styling -------------------------
    # For flat mode
    if mode == "flat":
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx, cell in enumerate(row[:-1]):
                if cell.value:
                    cell.fill = PatternFill(start_color=LEVEL_COLORS[col_idx % len(LEVEL_COLORS)],
                                            end_color=LEVEL_COLORS[col_idx % len(LEVEL_COLORS)],
                                            fill_type="solid")
                    cell.font = Font(bold=True)
            # Path - yellow
            row[-1].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    else:
        # merged / merged_full - merging cells
        for col_idx in range(max_level):
            current_val = None
            start_row = 2
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                val = cell.value
                if val:
                    color = LEVEL_COLORS[col_idx % len(LEVEL_COLORS)]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    if mode == "merged_full":
                        item_idx = row_idx - 2
                        if 0 <= item_idx < len(items):
                            cell.font = Font(bold=items[item_idx]["is_dir"])
                    else:
                        cell.font = Font(bold=True)
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(bold=False)

                if val != current_val:
                    if current_val is not None and start_row < row_idx - 1:
                        ws.merge_cells(start_row=start_row, start_column=col_idx + 1,
                                       end_row=row_idx - 1, end_column=col_idx + 1)
                    current_val = val
                    start_row = row_idx
            # Last group
            if current_val is not None and start_row < ws.max_row:
                ws.merge_cells(start_row=start_row, start_column=col_idx + 1,
                               end_row=ws.max_row, end_column=col_idx + 1)

        path_column = max_level + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=path_column, max_col=path_column):
            for cell in row:
                if cell.value:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --------------------- Common styles ----------------------
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.2, 50)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # --------------------- Statistics row ---------------------
    stat_row = ws.max_row + 2
    stat_text = f"Folders: {folder_count}, Files: {file_count}"
    last_col = get_column_letter(ws.max_column)
    ws.merge_cells(f"A{stat_row}:{last_col}{stat_row}")
    stat_cell = ws.cell(row=stat_row, column=1)
    stat_cell.value = stat_text
    stat_cell.fill = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")
    stat_cell.font = Font(bold=True, size=12)
    stat_cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_file)

# ----------------------------- Utility -----------------------------
def generate_output_path(input_file: str, mode: str) -> str:
    """Generates output filename in the same folder."""
    base = os.path.splitext(os.path.basename(input_file))[0]
    return os.path.join(os.path.dirname(os.path.abspath(input_file)), f"{base}_{mode}.xlsx")

# ------------------------------ Main -------------------------------
def main():
    parser = argparse.ArgumentParser(description="Convert tree output to Excel (simplified version)")
    parser.add_argument("-i", "--input", required=True, help="Path to input tree file")
    parser.add_argument("-m", "--mode", choices=["flat", "merged", "merged_full"],
                        default="flat", help="Output mode (default: flat)")
    parser.add_argument("-o", "--output", help="Path to output Excel (generated if not specified)")
    parser.add_argument("-e", "--encoding", help="Force encoding (e.g. utf-8, cp1251). Overrides auto-detection.")
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")

    args = parser.parse_args()

    # --------------- Check input file existence ---------------
    if not os.path.isfile(args.input):
        print(f"Error: file '{args.input}' not found.", file=sys.stderr)
        sys.exit(1)

    output_file = args.output if args.output else generate_output_path(args.input, args.mode)

    try:
        items, folders, files = parse_tree_file(args.input, args.encoding)
    except Exception as e:
        print(f"Parsing error: {e}", file=sys.stderr)
        sys.exit(1)

    if not items:
        print("Warning: no items found. Output file not created.", file=sys.stderr)
        sys.exit(0)

    try:
        save_to_excel(items, output_file, args.mode, folders, files)
    except Exception as e:
        print(f"Error saving Excel: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Done: {output_file}")

if __name__ == "__main__":
    main()
